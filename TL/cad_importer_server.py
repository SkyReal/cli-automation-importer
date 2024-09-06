# -*- coding: utf-8 -*-

from struct import pack
import argparse
import openpyxl
import socket 
from threading import Thread
from time import sleep
from select import select
import os 
import json 
from datetime import datetime
from subprocess import run
from wakeonlan import send_magic_packet
from struct import unpack 
from json import dump
from json import load 
from common import *

from getpass import getpass


IDLE_TIME_OUT = 1000     # seconds


# Mapping du reseau

def ask_password():
    password = getpass("Password : ")                     # saisir le mot de passe en masquant les caractères
    return password


def read_json(fichier_json):        # lit le fichier json et ajoute les adresses ip a celles pretes a travailler
    with open(fichier_json, 'r') as file:
        ip_dictionary = json.load(file)
    file.close()
    return ip_dictionary
    
def verif_excel_filename(excel_filename):
    excel_extensions = ['.xlsx']
    for extensions in  excel_extensions:
        if excel_filename.lower().endswith(extensions):
            return True 
    return False

def verif_repertory(CAD_repertory):   # verifier que notre deuxieme argument est valide 
    print("here the cad reportory :", CAD_repertory)
    if os.path.exists(CAD_repertory) == False:
        print('Either the path of your repertory does not exist, or you do not have permissions to access it ')
        return False 
    if os.path.isdir(CAD_repertory) == False and  os.path.isfile(CAD_repertory) == False: 
        return False
    print('Your repertory seems correct and will be scanned ')
    return True

def get_config_files_datas(path_CLI_local, JSON_file ):
    file = open(JSON_file, 'r')
    config_files_dictionnaire = json.load(file)                                     #on convertit le json en dictionnaire 
    path_CLI_local = config_files_dictionnaire["path_cli"] 
    if (path_CLI_local == None): 
        print(' ADRESSE_XRCENTER is missing in the config file.')
        return None
    return path_CLI_local
 
def replace_CAD_path_with_mapping(share_path, CAD_repertory, driver_letter):
    # Normalize path separators
    #share_path = os.path.normpath(share_path)
    #CAD_repertory = os.path.normpath(CAD_repertory)

    # Check if CAD_repertory starts with share_path
    if CAD_repertory.startswith(share_path):
        # If so replace it with the driver letter to scan using credentials
        new_path = CAD_repertory.replace(share_path, driver_letter + ":", 1)
        return new_path
    else:
        return CAD_repertory     

def clear_XRCENTER_config_file(base_address):
    xrcenter_config_file = "C:\\ProgramData\\Skydea\\xrcenter-cmd\\xrcenter-cmd.json"           # propre a skyreal
    with open(xrcenter_config_file, 'r') as xrcenter_file:
        xrcenter_dictionary = load(xrcenter_file)
    xrcenter_file.close()
    with open(xrcenter_config_file, 'w') as xrcenter_file_reopened:
        xrcenter_dictionary["XRCenter"]["BaseAddress"] = base_address
        dump(xrcenter_dictionary, xrcenter_file_reopened, indent=4)
    xrcenter_file_reopened.close()
    return

def verif_CLI(path_CLI_local):
    commande_CLI_opti1= f'& "{path_CLI_local}" health ping'
    CLI_ope = run(["Powershell", "-Command", commande_CLI_opti1], capture_output=True, text=True)
    if 'success' in CLI_ope.stdout.lower():                                #le XRCENTER se lance avec le chemin basique 
        print("CLI functional")
    else:                                                                       #on fait l'opération avec les paramètres que l'utilisateur a inséré
        print("Error : Is your CLI functional?")
        return False
    return True


def creation_workspace_deck(save_id_workspace, path_CLI_local):
    workspace_json={}
    print('A new workspace is created for your CAD files')
    date = datetime.now().strftime("%Y-%m-%d %H:%M")                    ## le nom du workspace correspond à la date de sa création 
    creation_workspace= f'& "{path_CLI_local}" workspace create --name workspace_"{date}"'
    nouveau_workspace= run(["Powershell", "-Command", creation_workspace], capture_output=True, text=True)
    workspace_json= json.loads(nouveau_workspace.stdout)
    save_id_workspace= workspace_json["EntityId"]
    if nouveau_workspace.stderr != '':   #on vérifie qu'il n'y a pas d'erreurs
        print('Problem with the workspace')
        return None
    return save_id_workspace



def find_files(file_list, CAD_repertory):   # fonction intermediaire
    for path, directory, files in os.walk(f"{CAD_repertory}"):
       for file_name in files:
            file_path = os.path.join(path, file_name)                   # Obtenir le chemin absolu du fichier
            file_list.append(file_path)                            # Ajouter le chemin absolu à la liste
            print("file found ", file_path)
    return 


def scan_CAD(file_list, CAD_files, CAD_repertory, extensions ):
    find_files(file_list, CAD_repertory)
    if file_list == [] :                                               # Le dossier est vide 
        print('This repertory is empty') 
        return False
    else: 
        for k in range(0,len(file_list)):
            for i in range(0, len(extensions)):                             #on s'arrête s'il n'existe pas de fichers CAD
                if file_list[k].lower().endswith(extensions[i].lower()) == True :      #c'est un dossier CAD, on regarde son extension en minuscule pour le rendre insensible a la casse
                    print("File added to import list : ", file_list[k])
                    if extensions[i].lower() == '.description':
                        handle_description_files( file_list[k], CAD_files  )
                        break
                    else :
                        CAD_files.append(file_list[k])
                        break                                                   #on n'a pas besoin de regarder le reste des extensions
    if (CAD_files == []):                                                # il n'y avait pas de fichiers à traiter
        print('There was no CAD files in this repertory')
        return False 
    return True


def create_excel(excel_filename):
    try:
        full_path =  os.path.join('C:\ProgramData\cli_automation_importer', excel_filename)
        workbook = openpyxl.load_workbook(full_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        default_sheet = workbook.active                                 # si le classeur est nouveau, on enleve la feuille par defaut
        workbook.remove(default_sheet)
    date = datetime.now().strftime("%y%m%d_%H%M")   
    sheet_name = f'Sheet_{date}'
    sheet = workbook.create_sheet(title=sheet_name)
    sheet.cell(row=1, column=1, value="CAD path")
    sheet.cell(row=1, column=2, value="Import Time")
    sheet.cell(row=1, column=3, value="Polygons")
    sheet.cell(row=1, column=4, value="Instances")
    sheet.cell(row=1, column=5, value="Prep")
    return sheet, workbook  
        
def results_in_excel( path, time,  sheet, excel_filename, workbook, polygons, instances, prep ):
    row_index = sheet.max_row + 1                        # on ecrit à la dernière colonne où il y a la place
    sheet.cell(row=row_index, column=1, value=path)
    sheet.cell(row=row_index, column=2, value=time)
    sheet.cell(row=row_index, column=3, value=polygons)
    sheet.cell(row=row_index, column=4, value=instances)
    sheet.cell(row=row_index, column=5, value=prep)
    
    full_path = os.path.join('C:\\ProgramData\\cli_automation_importer', excel_filename)
    workbook.save(full_path)
    print('Results have been successfully saved')
    return 


def check_CLI_version(path_CLI_local, CLI_version):
    powershell_command_cli_version = f'''
    & '{path_CLI_local}' --version
    '''
    powershell_check_CLI = run(["Powershell", "-Command", powershell_command_cli_version], capture_output=True, text=True)
    for caracters in powershell_check_CLI.stdout:
        if caracters == '-':
            break
        else: 
            CLI_version += caracters
    return CLI_version


def handle_description_files(file, cad_list):
    file_content = ''
    with open(file, 'r') as file_to_analyse:
        file_content = file_to_analyse.read()
    if 'xml' in file_content:
        return
    else: 
        try: 
            new_dictionary = json.loads(file_content)       # on fait devenir le .description, un fichier json en dictionnaire 
            product_name = new_dictionary["ProductFileName"] 
            new_path = os.path.dirname(file) + '\\' + product_name
            cad_list.append(new_path)
        except Exception:
            print(f'Your file {file} was not correct and will not be imported')
            return 
    return 
        
         
# PARTIE SERVEUR 

    
def get_informations(client_socket):
    try: 
        id_length = client_socket.recv(4)                                             # longeur de l'id workspace
        received_id_length = unpack('!I', id_length)[0]                               # traduit de binaire a string
        sleep(3)
        informations_bytes = client_socket.recv(received_id_length)                   # le fichier arrive en binaire, normalement sans erreur si il est apres select()        
        informations =informations_bytes.decode('utf-8')                             # on le retransforme en string
        sleep(3)
    except Exception:
        print('an error has occured, please try again')
        return ''
    return informations 

def send_wake_on_lan(client_socket, wake_on_lan):
    if wake_on_lan:
        message = 'activated'
    else: 
        message = 'desactivated'
    message_bytes=  message.encode('utf-8')
    message_length = pack('!I', len(message_bytes))
    client_socket.sendall(message_length + message_bytes)
    return 

def send_CLI_version(client_socket, CLI_version):
    CLI_version_bytes=  CLI_version.encode('utf-8')
    message_length = pack('!I', len(CLI_version_bytes))
    client_socket.sendall(message_length + CLI_version_bytes)
    return 


def send_workspace_id(client_socket, save_id_workspace):
    save_id_workspace_bytes=  save_id_workspace.encode('utf-8')
    message_length = pack('!I', len(save_id_workspace_bytes))
    client_socket.sendall(message_length + save_id_workspace_bytes)
    return 

def send_mapping_username(client_socket, mapping_username):
    mapping_username_bytes=  mapping_username.encode('utf-8')
    message_length = pack('!I', len(mapping_username_bytes))
    client_socket.sendall(message_length + mapping_username_bytes)
    return 
    
def send_mapping_password(client_socket, mapping_password):
    mapping_password_bytes=  mapping_password.encode('utf-8')
    message_length = pack('!I', len(mapping_password_bytes))
    client_socket.sendall(message_length + mapping_password_bytes)
    return 
    

def accept_connexions(serversocket, CAD_files, client_state, save_id_workspace, new_clients_counter, CAD_files_copy, mapping_username, mapping_password, CLI_version, wake_on_lan) :        
    while len(CAD_files) > 0 and new_clients_counter[0] < len(CAD_files_copy):           # condition d'arret du thread
        try:
            (new_client_socket, client_address) = serversocket.accept()   #pour accepter la connexion de clients  
            new_clients_counter[0] += 1
            send_wake_on_lan(new_client_socket, wake_on_lan)                # on regarde si le wake_on_lan est activé
            sleep(2)
            send_CLI_version(new_client_socket, CLI_version)
            sleep(2)
            send_mapping_username(new_client_socket, mapping_username)
            sleep(2)
            send_mapping_password(new_client_socket, mapping_password)
            sleep(2)
            send_workspace_id(new_client_socket, save_id_workspace)
            client_state[new_client_socket] = 'initialized'      #on initialise le premier client (sinon stack overflow)
            print(f' the workspace id was sent to {client_address[0]}')
            sleep(2)
        except socket.error as err:         #gerer les erreurs propres a la connexion des sockets 
            print(f"Socket error: {err}")
        except Exception as e:              # gerer les autres erreurs
            print(f"An unexpected error occurred while getting a new client: {e}")   
    return 
    
    

    
def computer_in_work(CAD_files, working_list, client_state, result_dictionary): 
    clients_list = client_state.keys()              # liste des elements du dictionnaire
    while True:
        if len(clients_list)>0:         # on verifie qu'il y en a au moins un a etudier
            for keys in clients_list:
                if client_state[keys] == 'initialized':
                    sleep(5)
                    client_state[keys] ='ready'
            clients_list_readable, client_list_writeable, client_error = select(clients_list, clients_list, clients_list, 10)    # methode non bloquante 
            if client_error != []:                                                                                               #gerer les erreurs
                for client_socket in client_error:
                    print(f'error with "{client_socket}", the client will now close')
                    for key in result_dictionary.keys():                                                                         #si il y a eu un probleme au niveau du client          
                        if result_dictionary[key] == client_socket:
                            result_dictionary[key] = -2
                    clients_list.remove(client_socket)                                                                           #on enleve le socket problematique 
                    client_socket.close()                                                                                        # on ferme le socket
            if client_list_writeable != []:                                                                                      # ces client sont prêt à recevoir des infos
                for client_sockets in client_list_writeable:
                    if len(CAD_files) >0 :
                        if client_state[client_sockets] == 'ready':
                            file_to_send = CAD_files.pop(0)                                                               # on prend le nouveau premier element de fichiers cad
                            file_to_send_bytes= file_to_send.encode('utf-8')                                                 # il faut envoyer le fichier en binaire
                            client_sockets.sendall(file_to_send_bytes)                                                       # on envoie le fichier au client dispo 
                            result_dictionary[file_to_send] = client_sockets                                                 # le path est en premier lieu associé au client
                            print(f'sending "{file_to_send}"')
                            client_state[client_sockets] = 'working'                                                         # ils ne sont plus pret
                            working_list.append(client_sockets)            
                            sleep(1)
        sleep(1)                                                                                                               # pas reelement necessaire mais c'est pour eviter un nombre trop important de boucles inutiles
    return 

    



def reception(working_list, client_state, CAD_files_copy, result_dictionary, number_of_received_import, excel_filename, sheet, workbook, list_of_informations):
    while True:
        if len(client_state)>0:
            if working_list != []:
                clients_reception_readable, clients_reception_writeable, client_reception_error = select(working_list, working_list, working_list, 5)
                if client_reception_error != []:
                    for client_socket in client_reception_error:
                        try:
                            print(f'error with "{client_socket}", the client will now close')
                            del client_state[client_socket]
                        except Exception:           # la cle a disparue?
                            print(f'error with "{client_socket}", the client can not be reach')         
                if clients_reception_readable != []:
                    for sleeping_clients in clients_reception_readable:
                        try:  
                            result_wait = get_informations(sleeping_clients)                                                      # on peut deja lire leurs contenus avec readable, il faudra juste gerer les erreurs de wifi
                            result_wait_float = float(result_wait)
                            polygons = get_informations(sleeping_clients)                                                      # on peut deja lire leurs contenus avec readable, il faudra juste gerer les erreurs de wifi
                            polygons_float = float(polygons)
                            instances = get_informations(sleeping_clients)                                                     
                            instances_float = float(instances)
                            result_prep = get_informations(sleeping_clients)                                                      
                            result_prep_float = float(result_prep)
                            for path in result_dictionary.keys():                                                               # on remplace les clients sockets par le timer recu apres l'import 
                                if result_dictionary[path] == sleeping_clients:
                                        result_dictionary[path] =  result_wait_float 
                                        list_of_informations[0] = polygons_float
                                        list_of_informations[1] = instances_float
                                        list_of_informations[2] = result_prep_float
                                        try:
                                            results_in_excel( path, result_wait_float,  sheet, excel_filename, workbook, list_of_informations[0] , list_of_informations[1], list_of_informations[2] )
                                        except KeyboardInterrupt:
                                            results_in_excel( path, result_wait_float,  sheet, excel_filename, workbook, 0, 0, 0 )
                                        break 
                            number_of_received_import[0] +=1
                            client_state[sleeping_clients] = 'ready'    
                            working_list.remove(sleeping_clients)
                            print('the result of the import was acquired')                            
                            print( number_of_received_import[0],'/', len(CAD_files_copy))
                        except socket.error:
                            print('Error while getting a client response')
                            client_state.pop(sleeping_clients)
                            for key in result_dictionary.keys():                                                               #si il y a eu un probleme de connexion          
                                if result_dictionary[key] == sleeping_clients:
                                    result_dictionary[key] = -2
                            number_of_received_import[0] +=1
                            working_list.remove(sleeping_clients)                                                              # on l'enleve completement de la liste
                            sleeping_clients.close()
                        except Exception:
                            client_state.pop(sleeping_clients)
                            for key in result_dictionary.keys():                                                               #si il y a eu un probleme de connexion          
                                if result_dictionary[key] == sleeping_clients:
                                    result_dictionary[key] = -2
                            number_of_received_import[0] +=1
                            working_list.remove(sleeping_clients)                                                              # on l'enleve completement de la liste
                            sleeping_clients.close()
                            
    return



def start_wake_on_lan(mac_addresses, wake_on_lan):
    if wake_on_lan:
        for key, values in mac_addresses.items():
            send_magic_packet(values)           # on reveille chacun des pc spécifiés
            sleep(180)                          # on attend 3 min entre chacun des pcs allumés  
    return 


def threading_in_progress(serversocket, CAD_files, working_list, client_state, save_id_workspace, CAD_files_copy, result_dictionary, number_of_received_import, new_clients_counter, mapping_username, mapping_password, CLI_version, mac_addresses, wake_on_lan, excel_filename,  sheet, workbook, list_of_informations):
    
    accept_thread = Thread(target = accept_connexions, args=(serversocket, CAD_files_copy, client_state, save_id_workspace, new_clients_counter, CAD_files_copy, mapping_username, mapping_password, CLI_version, wake_on_lan), daemon = True)          #on verifie en permanence si il y a une adresse dispo
    accept_thread.start()
    
    send_thread = Thread(target = computer_in_work, args= (CAD_files , working_list, client_state, result_dictionary), daemon = True)
    send_thread.start()    
    
    reception_thread = Thread(target = reception , args= (working_list, client_state, CAD_files_copy, result_dictionary, number_of_received_import, excel_filename,  sheet, workbook, list_of_informations), daemon = True)
    reception_thread.start()
    
    WOL_thread = Thread(target = start_wake_on_lan , args= (mac_addresses, wake_on_lan) , daemon = True)
    WOL_thread.start()
    
    return accept_thread,send_thread, reception_thread, WOL_thread
    

def closing_thread(accept_thread, send_thread,reception_thread, WOL_thread):
    accept_thread.join(0)    
    send_thread.join(0)
    reception_thread.join(0)
    WOL_thread.join(0)
    sleep(5)    
    return
    
def closing_clients(client_state):
    for kill_clients in client_state.keys():             # pour fermer tout les clients
        closing_message = 'close'                        
        closing_message_byte= closing_message.encode('utf-8')           #pour pouvoir l'envoyer
        kill_clients.sendall(closing_message_byte)      
    return
     

def main():  
    
    # informations sur les arguments de la fonction
    
    parser = argparse.ArgumentParser(description = 'Process the arguments')                     # va gerer les differents arguments
    parser.add_argument('--excel_filename', type = str, help = 'name of your excel file. If not specified, the results will be put in a file named "results_import_CAD"')
    parser.add_argument('--rep', type = str, help = 'name of your repertory containing the CAD files.', required= True)
    parser.add_argument('--credentials', type = str, help = 'file path of a json containing login and password fields')
    parser.add_argument('--json_mac', type = str, help = 'name of your json containing the mac adresses of the clients when the program ends.')
    parser.add_argument('-W', type = bool , help = 'enable WakeOnLan or disable it (True/False) ')
    parser.add_argument('--id_workspace', type = str , help = 'import in a specified workspace ')  #le nom du fichier excel


    args = parser.parse_args()  # parcourir les differents arguments
    
    excel_filename = args.excel_filename if args.excel_filename else 'results_import_CAD.xlsx'              # nom du fichier excel
    CAD_repertory =  args.rep                                                           # repertoire contenant les fichiers CAD
    JSON_file = "C:\ProgramData\cli_automation_importer\cad_importer_config_file.json"
    wake_on_lan = args.W if args.W else False                                           # de base le wake on lan n'est pas actif 
    json_with_mac = args.json_mac if args.json_mac else "C:\\ProgramData\\cli_automation_importer\\mac_addresses.json"
    save_id_workspace =  args.id_workspace if args.id_workspace else ''

    # vérification sur les arguments de la fonction
    
    if not verif_excel_filename(excel_filename):
        print('error : the extension of your excel filename must be : .xlsx' )
        return 
    
    if wake_on_lan:
        mac_addresses = read_json(json_with_mac)
    else:
        mac_addresses = ''

    # variables nécessaires pour la partie scan et infos
    
        # variable nécessaires au mapping 
       

    mapping_username, mapping_password = None, None

    if args.credentials:
        mapping_username, mapping_password = load_credentials_file(args.credentials)

    if mapping_username == None or mapping_password == None:
        mapping_username= input("username with domain name:")
        mapping_password = ask_password()

    share_path = ''
    drive_letter = 'Z'
    base_address = ''                   #   adresse du XRCenter de base avant le programme
    
        # scan 
        
    file_list=[]
    path_CLI_local = None  
    CLI_version = ''
    
    # initialisation des variables necessaires au serveur 
    
    global IDLE_TIME_OUT
    
    CAD_files = []
    working_list = []
    client_state = {}
    number_of_received_import = [0]           # pour compter le nombre d'import recu
    new_clients_counter = [0]
    list_of_informations = [0, 0, 0]   # for polygons, instances, and check if the 3d visualization was ok 
    
    # données du config file

    path_CLI_local, base_address, extensions_available, _ , share_path = load_config_file(JSON_file)
    
    # verif sur le mapping
    
    if not create_smb_mapping(mapping_username, mapping_password ,drive_letter, share_path):
        return
    
    global smb_mapped
    smb_mapped = True

    print("Old cad rep :", CAD_repertory)
    print("SHare path is :", share_path)
    CAD_repertory = replace_CAD_path_with_mapping(share_path, CAD_repertory, drive_letter)
    print("New cad rep :", CAD_repertory)

    # verifications sur les extensions 

    if extensions_available == []:
        print('your extensions are incorrect or not supported by Skyreal')
        return 
    
    # infos et vérifications sur la CLI 
    
    if  path_CLI_local== None:
        print("error while trying to read config file")
        return
    
    if not verif_CLI(path_CLI_local):            # on arrete le programme en cas d'erreur 
        return 
    
    CLI_version= check_CLI_version(path_CLI_local, CLI_version)
    
    # scan du dossier 
    
    if not verif_repertory(CAD_repertory):          # on verifie que notre repertoire est valide 
        return
    
    if not scan_CAD(file_list, CAD_files, CAD_repertory, extensions_available ) :
        return
    
    CAD_files_copy = list(CAD_files) # copie pour la condition
    result_dictionary = {path : 0 for path in CAD_files}         # a la fin, pour que chaque path soit associe a son temps d'import
    
    remove_smb_mapping(drive_letter)
    smb_mapped = False

    # creation du workspace et vérifications 
    if save_id_workspace == '':
        save_id_workspace= creation_workspace_deck(save_id_workspace, path_CLI_local)       # on va passer ce workspace aux clients
    
    if save_id_workspace== None  or save_id_workspace== '':                         # mais seulement si il est valable 
        print( 'error in the workspace creation')
        return
    else:
        print('your workspace id is :', save_id_workspace)
    
    # creation du serveur
    
    serversocket= socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    serversocket.bind(( socket.gethostbyname(socket.gethostname()), 3000) )        
    serversocket.listen()

    sheet, workbook = create_excel(excel_filename)
    
    
    try:
        accept_thread, send_thread, reception_thread, WOL_thread = threading_in_progress( serversocket, CAD_files, working_list, client_state, save_id_workspace, CAD_files_copy, result_dictionary, number_of_received_import, new_clients_counter, mapping_username, mapping_password, CLI_version, mac_addresses, wake_on_lan, excel_filename,  sheet, workbook, list_of_informations)  
    except KeyboardInterrupt:
        for clients in client_state.keys():
            clients.close()
        serversocket.close()

        
    while number_of_received_import[0] < len( CAD_files_copy ):
        if len(client_state) != 0:
            sleep(2)
        else: 
            print(' there are no clients...')
            sleep(5)
    
    closing_thread(accept_thread, send_thread, reception_thread, WOL_thread)
        
    closing_clients(client_state)   
    
    clear_XRCENTER_config_file(base_address)

    results_in_excel(result_dictionary, excel_filename, client_state)
    
    serversocket.close()
    
    return 
   
smb_mapped = False
main()       
if smb_mapped:
    remove_smb_mapping("Z")