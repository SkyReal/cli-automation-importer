# -*- coding: utf-8 -*-
"""
Created on Tue Jun 11 09:44:04 2024

@author: AxelBaillet
"""

#serveur host 

import argparse
import openpyxl
import socket 
from threading import Thread
from time import sleep
from select import select
import os 
import json 
from datetime import datetime
from subprocess import run
# extensions de fichiers CAD pris en compte par SkyReal

extensions = [".CATPart", ".CATProduct",".CGR","CATProcess",".model","3dxml",".plmxml",".jt", ".prt",".asm",".ifc",".sldprt",".sldasm",".stp", ".step",".stl",".iam",".ipt",".x_t",".dwg"]         
## certaines extensions sont en train d'être ajoutées : .fbx , .dgn . 

IDLE_TIME_OUT = 300     # temps en secondes pour lequel le programme s'arrete




#   PARTIE SCAN D'UN DOSSIER MIS EN ARGUMENT

# variables nécessaires pour la partie scan et infos


def verif_excel_filename(excel_filename):
    excel_extensions = ['.xlsx', '.xlsm', '.xltx','.xltm']
    for extensions in  excel_extensions:
        if excel_filename.lower().endswith(extensions):
            return True 
    return False





def verif_repertory(CAD_repertory):   # verifier que notre deuxieme argument est valide 
    if os.path.exists(CAD_repertory) == False:
        print('The path of your repertory does not exists')
        return False 
    if os.path.isdir(CAD_repertory) == False and  os.path.isfile(CAD_repertory) == False: 
        return False
    print('Your repertory seems correct and will be scanned ')
    return True


def get_config_files_datas(path_CLI_local, JSON_file ):
    if JSON_file.lower().endswith('.json'):          # on vérifie qu'on nous donne un json
        print('Your json file has been properly exploited')     
    else :
        print("This is not a json file")
        return None
    file = open(JSON_file, 'r')
    config_files_dictionnaire = json.load(file)                                     #on convertit le json en dictionnaire 
    path_CLI_local = config_files_dictionnaire["adresse_cli"] 
    if (path_CLI_local == None): 
        print(' ADRESSE_XRCENTER is missing in the config file.')
        return None
    return path_CLI_local


def verif_CLI(path_CLI_local):
    commande_CLI_opti1= f'& "{path_CLI_local}" health ping'
    CLI_ope = run(["Powershell", "-Command", commande_CLI_opti1], capture_output=True, text=True)
    if 'success' in CLI_ope.stdout.lower():                                #le XRCENTER se lance avec le chemin basique 
        print("CLI functional")
    else:                                                                       #on fait l'opération avec les paramètres que l'utilisateur a inséré
        print("Error : Is your CLI functional?")
        return False
    return True


def creation_workspace_deck(save_id_workspace, path_CLI_local):
    workspace_json={}
    print('A new workspace is created for your CAD files in your Deck')
    date = datetime.now().strftime("%Y-%m-%d %H:%M")                    ## le nom du workspace correspond à la date de sa création 
    creation_workspace= f'& "{path_CLI_local}" workspace create --name workspace_"{date}"'
    nouveau_workspace= run(["Powershell", "-Command", creation_workspace], capture_output=True, text=True)
    workspace_json= json.loads(nouveau_workspace.stdout)
    save_id_workspace= workspace_json["EntityId"]
    if nouveau_workspace.stderr != '':   #on vérifie qu'il n'y a pas d'erreurs
        print('Problem with the workspace')
        return None
    return save_id_workspace



def find_files(liste_fichiers, CAD_repertory):   # fonction intermediaire
    for path, directory, files in os.walk(f"{CAD_repertory}"):
       for file_name in files:
            file_path = os.path.join(path, file_name)                   # Obtenir le chemin absolu du fichier
            liste_fichiers.append(file_path)                            # Ajouter le chemin absolu à la liste
    return 


def scan_CAD(liste_fichiers, fichiers_CAD, CAD_repertory ):
    find_files(liste_fichiers, CAD_repertory)
    if liste_fichiers == [] :                                               # Le dossier est vide 
        print('The repertory is empty')
        return False
    else: 
        for k in range(0,len(liste_fichiers)):
            for i in range(0, len(extensions)):                             #on s'arrête s'il n'existe pas de fichers CAD
                if liste_fichiers[k].lower().endswith(extensions[i].lower()) == True :      #c'est un dossier CAD, on regarde son extension en minuscule pour le rendre insensible a la casse
                    fichiers_CAD.append(liste_fichiers[k])
                    break                                                   #on n'a pas besoin de regarder le reste des extensions
    if (fichiers_CAD == []):                                                # il n'y avait pas de fichiers à traiter
        print('There was no CAD files in this repertory')
        return False 
    return True


def results_in_excel( result_dictionary, excel_filename, client_state):
    if len(client_state) == 0:
            print('No results can be send in the excel file')
            return 
    else:
        try:
            workbook = openpyxl.load_workbook(excel_filename)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            default_sheet = workbook.active                                 # si le classeur est nouveau, on enleve la feuille par defaut
            workbook.remove(default_sheet)
        result_list = list(result_dictionary.items())
        sheet_name = f'Sheet_{len(workbook.sheetnames)}'
        sheet = workbook.create_sheet(title=sheet_name)
        for row_index, (key, value) in enumerate(result_list, start=1):
            sheet.cell(row=row_index, column=1, value=key)
            sheet.cell(row=row_index, column=2, value=value)
        workbook.save(excel_filename)
        print(f'Results have been successfully saved to {excel_filename}')
    return 


# PARTIE CREATION DU SERVEUR 

def send_workspace_id(client_socket, save_id_workspace):
    save_id_workspace_bytes=  save_id_workspace.encode('utf-8')
    client_socket.send(save_id_workspace_bytes)
    return 
    

def handle_clients(client_addresses, client_address, new_client_socket, client_state):
    if client_address[0] not in client_addresses:       # l'adresse ip du client n'est pas présente dans la liste
        client_addresses.add(client_address[0])
        client_state[new_client_socket] = 'ready'      #on initialise le premier client comme prêt a travailler
        print(f'the connexion with "{client_address[0]}" started' )
    else:
        if new_client_socket in client_state:
            print(f'the client "{client_address[0]}" is already connected, you can close this window') 
        else:
            print('there was a connexion error, your computer will try again to import CAD files')              # on va remettre tous les parametres a 0
            client_state[new_client_socket] = 'ready'
    return
    
def accept_connexions(serversocket, client_addresses, fichiers_CAD, client_state, save_id_workspace) :           # gerera les erreurs plus 
    while len(fichiers_CAD) > 0:           # condition d'arret du thread
        try:
            (new_client_socket, client_address) = serversocket.accept()   #pour accepter la connexion de clients    
            send_workspace_id(new_client_socket, save_id_workspace)
            handle_clients(client_addresses, client_address, new_client_socket, client_state)# on les repertorie 
            print(' the workspace id was sent')
            sleep(3)
        except socket.error as err:         #gerer les erreurs propres a la connexion des sockets 
            print(f"Socket error: {err}")
        except Exception as e:              # gerer les autres erreurs
            print(f"An unexpected error occurred while getting a new client: {e}")   
    return 
    

def computer_in_work(fichiers_CAD, working_list, client_state, result_dictionary): 
    clients_list = client_state.keys()              # liste des elements du dictionnaire
    while True:
        if len(clients_list)>0:         # on verifie qu'il y en a au moins un a etudier
            clients_list_readable, client_list_writeable, client_error = select(clients_list, clients_list, clients_list, 10)    # methode non bloquante 
            if client_error != []:                                                                                               #gerer les erreurs
                for client_socket in client_error:
                    print(f'error with "{client_socket}", the client will now close')
                    for key in result_dictionary.keys():                                                                         #si il y a eu un probleme au niveau du client          
                        if result_dictionary[key] == client_socket:
                            result_dictionary[key] = -2
                    clients_list.remove(client_socket)                                                                           #on enleve le socket problematique 
                    client_socket.close()                                                                                        # on ferme le socket
            if client_list_writeable != []:                                                                                      # ces client sont prêt à recevoir des infos
                    for client_sockets in client_list_writeable:
                        if len(fichiers_CAD) >0 :
                            if client_state[client_sockets] == 'ready':
                                file_to_send = fichiers_CAD.pop(0)                                                               # on prend le nouveau premier element de fichiers cad
                                file_to_send_bytes= file_to_send.encode('utf-8')                                                 # il faut envoyer le fichier en binaire
                                client_sockets.sendall(file_to_send_bytes)                                                       # on envoie le fichier au client dispo 
                                result_dictionary[file_to_send] = client_sockets                                                 # le path est en premier lieu associé au client
                                print(f'sending "{file_to_send}"')
                                client_state[client_sockets] = 'working'                                                         # ils ne sont plus pret
                                working_list.append(client_sockets)            
                                sleep(1)
        sleep(0.5)                                                                                                               # pas reelement necessaire mais c'est pour eviter un nombre trop important de boucles inutiles
    return 

    



def reception(working_list, client_state, fichiers_CAD_copy, result_dictionary, number_of_received_import):
    while True:
        if len(client_state)>0:
            if working_list != []:
                clients_reception_readable, clients_reception_writeable, client_reception_error = select(working_list, working_list, working_list, 5)
                if client_reception_error != []:
                    for client_socket in client_reception_error:
                        try:
                            print(f'error with "{client_socket}", the client will now close')
                            del client_state[client_socket]
                        except Exception:           # la cle a disparue?
                            print(f'error with "{client_socket}", the client can not be reach')         
                if clients_reception_readable != []:
                    for sleeping_clients in clients_reception_readable:
                        try:                                                                                                    # on peut deja lire leurs contenus avec readable, il faudra juste gerer les erreurs de wifi
                            result_wait_bytes = sleeping_clients.recv(1024)                                                     # on regarde si ils renvoient quelque chose
                            result_wait = result_wait_bytes.decode()
                            result_wait_float = float(result_wait)
                            for path in result_dictionary.keys():                                                               # on remplace les clients sockets par le timer recu apres l'import 
                                if result_dictionary[path] == sleeping_clients:
                                        result_dictionary[path] =  result_wait_float
                            number_of_received_import[0] +=1
                            client_state[sleeping_clients] = 'ready'    
                            working_list.remove(sleeping_clients)
                            print('the result of the import was acquired')
                            print( number_of_received_import[0],'/', len(fichiers_CAD_copy))
                        except socket.error:
                            print('Error while getting a client response')
                            client_state.pop(sleeping_clients)
                            for key in result_dictionary.keys():                                                               #si il y a eu un probleme de connexion          
                                if result_dictionary[key] == sleeping_clients:
                                    result_dictionary[key] = -2
                            number_of_received_import[0] +=1
                            working_list.remove(sleeping_clients)            # on l'enleve completement de la liste
                            sleeping_clients.close()
                        except Exception:
                            client_state.pop(sleeping_clients)
                            for key in result_dictionary.keys():                                                               #si il y a eu un probleme de connexion          
                                if result_dictionary[key] == sleeping_clients:
                                    result_dictionary[key] = -2
                            number_of_received_import[0] +=1
                            working_list.remove(sleeping_clients)            # on l'enleve completement de la liste
                            sleeping_clients.close()
                            
    return


def threading_in_progress(serversocket, client_addresses, fichiers_CAD, working_list, client_state, save_id_workspace, fichiers_CAD_copy, result_dictionary, number_of_received_import):
    
    accept_thread = Thread(target = accept_connexions, args=(serversocket, client_addresses, fichiers_CAD, client_state, save_id_workspace), daemon = True)          #on verifie en permanence si il y a une adresse dispo
    accept_thread.start()
    
    send_thread = Thread(target = computer_in_work, args= (fichiers_CAD , working_list, client_state, result_dictionary), daemon = True )
    send_thread.start()    
    
    reception_thread = Thread(target = reception , args= (working_list, client_state, fichiers_CAD_copy, result_dictionary, number_of_received_import), daemon = True )
    reception_thread.start()
    
    return accept_thread,send_thread, reception_thread
    

def closing_thread(accept_thread, send_thread,reception_thread):
    accept_thread.join(0)    
    send_thread.join(0)
    reception_thread.join(0)
    sleep(5)    
    return
    
def closing_clients(client_state):
    for kill_clients in client_state.keys():             # pour fermer tout les clients
        closing_message = 'close'                        
        closing_message_byte= closing_message.encode('utf-8')           #pour pouvoir l'envoyer
        kill_clients.sendall(closing_message_byte)      
    return
     

def main():  
    
    
    # informations sur les arguments de la fonction
    
    parser = argparse.ArgumentParser(description = 'Process the three arguments')                     # va gerer les differents arguments
    parser.add_argument('--name', type = str, help = 'name of your excel file. If not specified, the results will be put in a file name "results_import_CAD"')  #le nom du fichier excel
    parser.add_argument('--rep', type = str, help = 'name of your repertory containing the CAD files.', required= True)
    parser.add_argument('--json', type = str, help = 'name of your json containing the informations required to make the program work.', required= True)

    args = parser.parse_args()  # parcourir les differents arguments
    
    
    excel_filename = args.name if args.name else 'results_import_CAD.xlsx'              # nom du fichier excel
    CAD_repertory =  args.rep                                                           # repertoire contenant les fichiers CAD
    JSON_file = args.json 
    
    
    # vérification sur les arguments de la fonction
    
    if not verif_repertory(CAD_repertory):
        return
    
    if not verif_excel_filename(excel_filename):
        print('error : the extension of your excel filename must belong the the following ones : .xlsx, .xlsm, .xltx,.xltm' )
        return 
    # variables nécessaires pour la partie scan et infos
    
    liste_fichiers=[]
    path_CLI_local = None
    save_id_workspace=''     
    
    # initialisation des variables necessaires au serveur 
    
    global IDLE_TIME_OUT
    
    fichiers_CAD = []
    client_addresses=  set()  #ensemble non ordonnées d'elements, on va s'en servir pour être sur qu'un serveur client ne s'enregistre pas 2 fois 
    working_list = []
    client_state = {}
    timer = 0
    number_of_received_import = [0]           # pour compter le nombre d'import recu
    
    
    
    
    
    
    # infos et vérifications sur la CLI 
    
    path_CLI_local= get_config_files_datas(path_CLI_local, JSON_file)
    if  path_CLI_local== None:
        print("error while trying to read config file")
        return
    
    if not verif_CLI(path_CLI_local):            # on arrete le programme en cas d'erreur 
        return 
    
    # creation du workspace et vérifications 
    
    save_id_workspace= creation_workspace_deck(save_id_workspace, path_CLI_local)       # on va passer ce workspace aux clients
    
    if save_id_workspace== None  or save_id_workspace== '':                         # mais seulement si il est valable 
        print( 'error in the workspace creation')
        return
    
    
    # scan du dossier 
    
    if not scan_CAD(liste_fichiers, fichiers_CAD, CAD_repertory) :
        return
    
    fichiers_CAD_copy = list(fichiers_CAD) # copie pour la condition
    result_dictionary = {path : 0 for path in fichiers_CAD}         # a la fin, pour que chaque path soit associe a son temps d'import
    
    # creation du serveur
    
    serversocket= socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    serversocket.bind(( socket.gethostbyname(socket.gethostname()), 3000) )        
    serversocket.listen()
    print('host server is starting \nwaiting for a first client')
   
    # il faut au moins un client pour pouvoir demarrer le script 
     
    # if manage_first_client(serversocket, client_state, save_id_workspace) == False:
    #     return 
        
    # execution du serveur
    
    try:
        accept_thread, send_thread, reception_thread = threading_in_progress( serversocket, client_addresses, fichiers_CAD, working_list, client_state, save_id_workspace, fichiers_CAD_copy, result_dictionary, number_of_received_import)  
    except KeyboardInterrupt:
        for clients in client_state.keys():
            clients.close()
            serversocket.close()
    
    while timer < IDLE_TIME_OUT and number_of_received_import[0] < len( fichiers_CAD_copy ):
        if len(client_state) != 0:
            timer = 0  
            sleep(2)
        else: 
            print(' there are no clients, remaining time before disconnexion',  IDLE_TIME_OUT - timer)
            sleep(5)
            timer += 5
    
    closing_thread(accept_thread, send_thread, reception_thread)
    
    results_in_excel( result_dictionary, excel_filename, client_state )
    
    closing_clients(client_state)
    
    serversocket.close()
    return 
   
    
main()       









