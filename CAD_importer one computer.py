# -*- coding: utf-8 -*-
"""
Created on Tue Jun  4 11:55:57 2024

@author: AxelBaillet
"""

## penser a  lancer python en ADMIN 
##cela ne marchera qu'en local pour le moment 
# les paths CLI par défaut sont ils vbraiment ceux par défaut?
## bibliothèques

from json import dump
from subprocess import run 
from json import load
import argparse
import os 
import subprocess
from datetime import datetime
from time import time 
import json 
import openpyxl

extensions = [".CATPart", ".CATProduct",".CGR","CATProcess",".model","3dxml",".plmxml",".jt", ".prt",".asm",".ifc",".sldprt",".sldasm",".stp", ".step",".stl",".iam",".ipt",".x_t",".dwg"]
## certaines extensions sont en train d'être ajoutées : .fbx , .dgn . 


# prendre le bon XRCENTER

def always_more_config_file_datas(JSON_file): 
    xrcenter_config_file = "C:\\ProgramData\\Skydea\\xrcenter-cmd\\xrcenter-cmd.json"           # propre a skyreal
    try:
        with open(JSON_file, 'r') as file: 
            config_files_dictionary = load(file)   
            ip_address_XRCENTER = config_files_dictionary["ip_address_XRCENTER"]  
            file.close()
    except Exception as e:
        print(e)
        return ''
    try:
        with open(xrcenter_config_file, 'r') as xrcenter_file:
            xrcenter_dictionary = load(xrcenter_file)
            base_address = xrcenter_dictionary["XRCenter"]["BaseAddress"]
            new_base_address = f'https://{ip_address_XRCENTER}:9228/'
        xrcenter_file.close()
        with open(xrcenter_config_file, 'w') as xrcenter_file_reopened:
            xrcenter_dictionary["XRCenter"]["BaseAddress"] = new_base_address
            dump(xrcenter_dictionary, xrcenter_file_reopened, indent=4)
        xrcenter_file_reopened.close()
    except Exception as e:
        print(e)
        return ''
    return base_address


def clear_XRCENTER_config_file(base_address):
    xrcenter_config_file = "C:\\ProgramData\\Skydea\\xrcenter-cmd\\xrcenter-cmd.json"           # propre a skyreal
    with open(xrcenter_config_file, 'r') as xrcenter_file:
        xrcenter_dictionary = load(xrcenter_file)
    xrcenter_file.close()
    with open(xrcenter_config_file, 'w') as xrcenter_file_reopened:
        xrcenter_dictionary["XRCenter"]["BaseAddress"] = base_address
        dump(xrcenter_dictionary, xrcenter_file_reopened, indent=4)
    xrcenter_file_reopened.close()
    return


# verifier le repertory 

def verif_repertory(CAD_repertory):   # verifier que notre deuxieme argument est valide 
    if os.path.exists(CAD_repertory) == False:
        print('Either the path of your repertory does not exist, or you do not have permissions to access it ')
        return False 
    if os.path.isdir(CAD_repertory) == False and  os.path.isfile(CAD_repertory) == False: 
        return False
    print('Your repertory seems correct and will be scanned ')
    return True


# Il faut un config files pour faire marcher le programme

    
def get_config_files_datas(path_CLI_local, JSON_file ):
    if JSON_file.lower().endswith('.json'):          # on vérifie qu'on nous donne un json
        pass 
    else :
        print("The config file is not a json file")
        return None
    with open(JSON_file, 'r') as file:
        config_files_dictionnaire = load(file)                                     #on convertit le json en dictionnaire 
        path_CLI_local = config_files_dictionnaire["path_cli"] 
        if (path_CLI_local == None): 
            print(' adress_xrcenter is missing in the config file.')
            return None
    file.close()
    return path_CLI_local
        
def verif_CLI(path_CLI_local):
    x=0
    if not os.path.exists(path_CLI_local):
        print('You need to write the correct local path of your cli in the config file')
        return False
    commande_CLI_opti1= f'& "{path_CLI_local}" health ping'
    while x < 10:                                                                                           # max 10 essais
        CLI_ope = run(["Powershell", "-Command", commande_CLI_opti1], capture_output=True, text=True)
        if 'success' in CLI_ope.stdout.lower():                                                             #le XRCENTER se lance avec le chemin basique 
            print("CLI functional")
            return True 
        else:                                                                                               #on fait l'opération avec les paramètres que l'utilisateur a inséré
            print("Trying to join the XRCenter")
            stdout = CLI_ope.stdout
            print(stdout)
            x += 1                  
    return False


## créer un nouveau workspace

def creation_workspace_deck(save_id_workspace, path_CLI_local):
    workspace_json={}
    print('A new workspace is created for your CAD files')
    date = datetime.now().strftime("%Y-%m-%d %H:%M")                    ## le nom du workspace correspond à la date de sa création 
    creation_workspace= f'& "{path_CLI_local}" workspace create --name workspace_"{date}"'
    nouveau_workspace= run(["Powershell", "-Command", creation_workspace], capture_output=True, text=True)
    workspace_json= json.loads(nouveau_workspace.stdout)
    save_id_workspace= workspace_json["EntityId"]
    if nouveau_workspace.stderr != '':   #on vérifie qu'il n'y a pas d'erreurs
        print('Problem with the workspace')
        return None
    return save_id_workspace


## Chercher les paths des fichiers du dossier passé en argument



def find_files(file_list, CAD_repertory):   # fonction intermediaire
    for path, directory, files in os.walk(f"{CAD_repertory}"):
       for file_name in files:
            file_path = os.path.join(path, file_name)                   # Obtenir le chemin absolu du fichier
            file_list.append(file_path)                            # Ajouter le chemin absolu à la liste
    return 



## scanner ces dossiers voir si il existe des fichiers cad 


def scan_CAD(file_list, CAD_files, CAD_repertory):
    find_files(file_list, CAD_repertory)
    if file_list == [] :                                               # Le dossier est vide 
        print('The repertory is empty')
        return []
    else: 
        for k in range(0,len(file_list)):
            for i in range(0, len(extensions)):                             #on s'arrête s'il n'existe pas de fichers CAD
                if file_list[k].lower().endswith(extensions[i].lower()) == True :      #c'est un dossier CAD, on regarde son extension en minuscule pour le rendre insensible a la casse
                    CAD_files.append(file_list[k])
                    break                                                   #on n'a pas besoin de regarder le reste des extensions
    if (CAD_files == []):                                                # il n'y avait pas de fichiers à traiter
        print('No file needed to be processed')
        return [] 
    return CAD_files


    

# Importer les fichiers scannés dans le CLI     



def import_fichier_CAD(time_record, save_id_workspace, path_CLI_local, fichier):
    print(f'The file {fichier} is send in SkyReal')
    start = time()
    commande_fichier_import= f'& "{path_CLI_local}"  cad import "{save_id_workspace}" "{fichier}"' #commande pour importer sur le deck
    import_final = subprocess.run(["Powershell", "-Command", commande_fichier_import], capture_output=True, text=True)
    end = time()
    temps_import= end - start
    if "failed" in import_final.stdout.lower() or "error" in import_final.stdout.lower():
            print(f'the file "{fichier}" was not imported (error)')   
            time_record.append('Error')
    else: 
        time_record.append(temps_import)                        #on mesure le temps de chaque import
    return
    

# importer le dossier complet
    
def import_dossiers_CAD(CAD_files, save_id_workspace , path_CLI_local, time_record):
    print("The import of your CAD files begins")
    for k in range(0,len(CAD_files)):
        import_fichier_CAD(time_record, save_id_workspace, path_CLI_local, CAD_files[k])
        print(f' {k+1} / {len(CAD_files)} ')
    return True



def write_in_excel(CAD_files, time_record, excel_filename):  #fichier .xlsx
    try:
        full_path =  os.path.join('C:\ProgramData\cli_automation_importer', excel_filename)
        workbook = openpyxl.load_workbook(full_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        default_sheet = workbook.active                                 # si le classeur est nouveau, on enleve la feuille par defaut
        workbook.remove(default_sheet)   
    data_in_excel = []
    for k in range(0, len(time_record)):
        data_in_excel.append([CAD_files[k], time_record[k]])
    workbook = openpyxl.Workbook()          # créer un nouveau workbook 
    sheet_name = f'Sheet_{len(workbook.sheetnames)}'
    sheet = workbook.create_sheet(title=sheet_name)
    for row_index, row_data in enumerate(data_in_excel, start=1):        #on se balade dans les lignes à partir de la première
        for col_index, cell_data in enumerate(row_data, start=1):          #on se balade dans les colonnes à partir de la première
            sheet.cell(row=row_index, column=col_index, value=cell_data)            
    workbook.save(full_path)   # on sauvegarde les données
    print(f"The datas were saved in '{excel_filename}'")
    return data_in_excel


   



def main():
    
    parser = argparse.ArgumentParser(description = 'Process the arguments')                     # va gerer les differents arguments
    parser.add_argument('--rep', type = str, help = 'name of your repertory containing the CAD files.', required= True)
    parser.add_argument('--excel_filename', type = str, help = 'name of your excel file. If not specified, the results will be put in a file named "results_import_CAD"')  #le nom du fichier excel

    args = parser.parse_args()  # parcourir les differents arguments

    excel_filename = args.excel_filename if args.excel_filename else 'results_import_CAD.xlsx'              # nom du fichier excel  
    CAD_repertory =  args.rep     
    JSON_file = "C:\ProgramData\cli_automation_importer\cad_importer_config_file.json"
    path_CLI_local = None
    base_address = always_more_config_file_datas(JSON_file)
    
    # scan du dossier 
  
    if not verif_repertory(CAD_repertory):          # on verifie que notre repertoire est valide 
          return
    
    
    path_CLI_local = get_config_files_datas(path_CLI_local, JSON_file )
    
    if  path_CLI_local== None:
        print("error while trying to read config file")
        return
    
    if not verif_CLI(path_CLI_local):            # on arrete le programme en cas d'erreur 
        return 

    
    file_list=[]
    CAD_files = []
    save_id_workspace=''                    
    time_record= []
    
    save_id_workspace= creation_workspace_deck(save_id_workspace, path_CLI_local)
    
    if save_id_workspace== None  or save_id_workspace== '':
        print( 'error in the workspace creation')
        return
    
    CAD_files = scan_CAD(file_list, CAD_files, CAD_repertory) 
    
    if CAD_files == []:
        return 
    
    try: 
        import_dossiers_CAD(CAD_files, save_id_workspace, path_CLI_local, time_record)
    except KeyboardInterrupt:
        pass
    finally:
        write_in_excel(CAD_files, time_record, excel_filename)
        clear_XRCENTER_config_file(base_address)
    
    return 
    
main()   ## on execute la fonction 
    
    
    
    
    
    
 




