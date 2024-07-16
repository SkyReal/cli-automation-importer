# -*- coding: utf-8 -*-
"""
Created on Tue Jun 25 10:39:50 2024

@author: AxelBaillet
"""

#serveur host 

from struct import pack
import argparse
import openpyxl
import socket 
from threading import Thread
from time import sleep
from select import select
import os 
import json 
from datetime import datetime
from subprocess import run
from wakeonlan import send_magic_packet

from getpass import getpass

# extensions de fichiers CAD pris en compte par SkyReal

extensions = [".CATPart", ".CATProduct",".CGR","CATProcess",".model","3dxml",".plmxml",".jt", ".prt",".asm",".ifc",".sldprt",".sldasm",".stp", ".step",".stl",".iam",".ipt",".x_t",".dwg"]         
## certaines extensions sont en train d'être ajoutées : .fbx , .dgn . 

IDLE_TIME_OUT = 1000     # temps en secondes pour lequel le programme s'arrete




# Mapping du reseau

def create_mapping_road(username, password, drive_letter, share_path):
    
    powershell_identification_orders= f"""
    $user = '{username}'
    $securePassword = '{password}'
    $pass = ConvertTo-SecureString $securePassword -AsPlainText -Force
    $credential = New-Object System.Management.Automation.PSCredential ($user, $pass)
    Write-Output "credentials"
    $credential
    New-SMbGlobalMapping -RemotePath '{share_path}' -Credential $credential -LocalPath {drive_letter}
    """
    try:
        powershell_command_1 = run([ "Powershell", "-Command",  powershell_identification_orders], capture_output=True, text=True)
        print('This program mapped the repertory you asked for on the letter', drive_letter)
    except Exception as e:
        print("error :", e)
        print(powershell_command_1.stderr)
        return False
    return True

def correct_mapping_letter(drive_letter):
    possible_letters = ['Z', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y']
    for letter in possible_letters:
        powershell_verification_orders = "Get-PSDrive -PSProvider FileSystem | Select-Object -ExpandProperty Name"
        powershell_command_verif = run([ "Powershell", "-Command",  powershell_verification_orders], capture_output=True, text=True)
        if letter not in powershell_command_verif.stdout:
            drive_letter = letter
    return  drive_letter


def ask_password():
    password = getpass("Password : ")                     # saisir le mot de passe en masquant les caractères
    return password

# nouveautes wake on lan

def computer_wake_up(ip_address):           # marche seulement si le pc est eteint 
    send_magic_packet('FF:FF:FF:FF:FF:FF', ip_address=ip_address)   # il faut que les adresses ip soient fixes
    sleep(50)
    return


def read_json(fichier_json):        # lit le fichier json et ajoute les adresses ip a celles pretes a travailler
    with open('fichier_json', 'r') as j:
        ip_list = json.load(j)
    return ip_list
    

#   PARTIE SCAN D'UN DOSSIER MIS EN ARGUMENT

def verif_excel_filename(excel_filename):
    excel_extensions = ['.xlsx', '.xlsm', '.xltx','.xltm']
    for extensions in  excel_extensions:
        if excel_filename.lower().endswith(extensions):
            return True 
    return False


def verif_repertory(CAD_repertory):   # verifier que notre deuxieme argument est valide 
    if os.path.exists(CAD_repertory) == False:
        print('Either the path of your repertory does not exist, or you do not have permissions to access it ')
        return False 
    if os.path.isdir(CAD_repertory) == False and  os.path.isfile(CAD_repertory) == False: 
        return False
    print('Your repertory seems correct and will be scanned ')
    return True

def get_config_files_datas(path_CLI_local, JSON_file ):
    if JSON_file.lower().endswith('.json'):          # on vérifie qu'on nous donne un json
        print('Your json file has been properly exploited')     
    else :
        print(" Your json file is not a json")
        return None
    file = open(JSON_file, 'r')
    config_files_dictionnaire = json.load(file)                                     #on convertit le json en dictionnaire 
    path_CLI_local = config_files_dictionnaire["adresse_cli"] 
    if (path_CLI_local == None): 
        print(' ADRESSE_XRCENTER is missing in the config file.')
        return None
    return path_CLI_local


def verif_CLI(path_CLI_local):
    commande_CLI_opti1= f'& "{path_CLI_local}" health ping'
    CLI_ope = run(["Powershell", "-Command", commande_CLI_opti1], capture_output=True, text=True)
    if 'success' in CLI_ope.stdout.lower():                                #le XRCENTER se lance avec le chemin basique 
        print("CLI functional")
    else:                                                                       #on fait l'opération avec les paramètres que l'utilisateur a inséré
        print("Error : Is your CLI functional?")
        return False
    return True


def creation_workspace_deck(save_id_workspace, path_CLI_local):
    workspace_json={}
    print('A new workspace is created for your CAD files')
    date = datetime.now().strftime("%Y-%m-%d %H:%M")                    ## le nom du workspace correspond à la date de sa création 
    creation_workspace= f'& "{path_CLI_local}" workspace create --name workspace_"{date}"'
    nouveau_workspace= run(["Powershell", "-Command", creation_workspace], capture_output=True, text=True)
    workspace_json= json.loads(nouveau_workspace.stdout)
    save_id_workspace= workspace_json["EntityId"]
    if nouveau_workspace.stderr != '':   #on vérifie qu'il n'y a pas d'erreurs
        print('Problem with the workspace')
        return None
    return save_id_workspace



def find_files(file_list, CAD_repertory):   # fonction intermediaire
    for path, directory, files in os.walk(f"{CAD_repertory}"):
       for file_name in files:
            file_path = os.path.join(path, file_name)                   # Obtenir le chemin absolu du fichier
            file_list.append(file_path)                            # Ajouter le chemin absolu à la liste
    return 


def scan_CAD(file_list, fichiers_CAD, CAD_repertory ):
    find_files(file_list, CAD_repertory)
    if file_list == [] :                                               # Le dossier est vide 
        print('This repertory is empty') 
        return False
    else: 
        for k in range(0,len(file_list)):
            for i in range(0, len(extensions)):                             #on s'arrête s'il n'existe pas de fichers CAD
                if file_list[k].lower().endswith(extensions[i].lower()) == True :      #c'est un dossier CAD, on regarde son extension en minuscule pour le rendre insensible a la casse
                    fichiers_CAD.append(file_list[k])
                    break                                                   #on n'a pas besoin de regarder le reste des extensions
    if (fichiers_CAD == []):                                                # il n'y avait pas de fichiers à traiter
        print('There was no CAD files in this repertory')
        return False 
    return True


def results_in_excel( result_dictionary, excel_filename, client_state):
    if len(client_state) == 0:
            print('No results can be send in the excel file')
            return 
    else:
        try:
            full_path =  os.path.join('C:\ProgramData\cli_automation_importer', excel_filename)
            workbook = openpyxl.load_workbook(full_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            default_sheet = workbook.active                                 # si le classeur est nouveau, on enleve la feuille par defaut
            workbook.remove(default_sheet)
        result_list = list(result_dictionary.items())
        sheet_name = f'Sheet_{len(workbook.sheetnames)}'
        sheet = workbook.create_sheet(title=sheet_name)
        for row_index, (key, value) in enumerate(result_list, start=1):
            sheet.cell(row=row_index, column=1, value=key)
            sheet.cell(row=row_index, column=2, value=value)
        workbook.save(full_path)
        print(f'Results have been successfully saved to {excel_filename}')
    return 


def check_CLI_version(path_CLI_local, CLI_version):
    powershell_command_cli_version = f'''
    & '{path_CLI_local}' --version
    '''
    powershell_check_CLI = run(["Powershell", "-Command", powershell_command_cli_version], capture_output=True, text=True)
    for caracters in powershell_check_CLI.stdout:
        if caracters == '-':
            break
        else: 
            CLI_version += caracters
    return CLI_version


# PARTIE SERVEUR 

def send_CLI_version(client_socket, CLI_version):
    CLI_version_bytes=  CLI_version.encode('utf-8')
    message_length = pack('!I', len(CLI_version_bytes))
    client_socket.sendall(message_length + CLI_version_bytes)
    return 


def send_workspace_id(client_socket, save_id_workspace):
    save_id_workspace_bytes=  save_id_workspace.encode('utf-8')
    message_length = pack('!I', len(save_id_workspace_bytes))
    client_socket.sendall(message_length + save_id_workspace_bytes)
    return 

def send_mapping_username(client_socket, mapping_username):
    mapping_username_bytes=  mapping_username.encode('utf-8')
    message_length = pack('!I', len(mapping_username_bytes))
    client_socket.sendall(message_length + mapping_username_bytes)
    return 
    
def send_mapping_password(client_socket, mapping_password):
    mapping_password_bytes=  mapping_password.encode('utf-8')
    message_length = pack('!I', len(mapping_password_bytes))
    client_socket.sendall(message_length + mapping_password_bytes)
    return 
    

def accept_connexions(serversocket, fichiers_CAD, client_state, save_id_workspace, new_clients_counter, fichiers_CAD_copy, mapping_username, mapping_password, CLI_version) :        
    while len(fichiers_CAD) > 0 and new_clients_counter[0] < len(fichiers_CAD_copy):           # condition d'arret du thread
        try:
            (new_client_socket, client_address) = serversocket.accept()   #pour accepter la connexion de clients  
            new_clients_counter[0] += 1
            send_CLI_version(new_client_socket, CLI_version)
            sleep(2)
            send_mapping_username(new_client_socket, mapping_username)
            sleep(2)
            send_mapping_password(new_client_socket, mapping_password)
            sleep(2)
            send_workspace_id(new_client_socket, save_id_workspace)
            client_state[new_client_socket] = 'initialized'      #on initialise le premier client (sinon stack overflow)
            print(f' the workspace id was sent to {client_address[0]}')
            sleep(2)
        except socket.error as err:         #gerer les erreurs propres a la connexion des sockets 
            print(f"Socket error: {err}")
        except Exception as e:              # gerer les autres erreurs
            print(f"An unexpected error occurred while getting a new client: {e}")   
    return 
    
    

    
def computer_in_work(fichiers_CAD, working_list, client_state, result_dictionary): 
    clients_list = client_state.keys()              # liste des elements du dictionnaire
    while True:
        if len(clients_list)>0:         # on verifie qu'il y en a au moins un a etudier
            for keys in clients_list:
                if client_state[keys] == 'initialized':
                    sleep(5)
                    client_state[keys] ='ready'
            clients_list_readable, client_list_writeable, client_error = select(clients_list, clients_list, clients_list, 10)    # methode non bloquante 
            if client_error != []:                                                                                               #gerer les erreurs
                for client_socket in client_error:
                    print(f'error with "{client_socket}", the client will now close')
                    for key in result_dictionary.keys():                                                                         #si il y a eu un probleme au niveau du client          
                        if result_dictionary[key] == client_socket:
                            result_dictionary[key] = -2
                    clients_list.remove(client_socket)                                                                           #on enleve le socket problematique 
                    client_socket.close()                                                                                        # on ferme le socket
            if client_list_writeable != []:                                                                                      # ces client sont prêt à recevoir des infos
                for client_sockets in client_list_writeable:
                    if len(fichiers_CAD) >0 :
                        if client_state[client_sockets] == 'ready':
                            file_to_send = fichiers_CAD.pop(0)                                                               # on prend le nouveau premier element de fichiers cad
                            file_to_send_bytes= file_to_send.encode('utf-8')                                                 # il faut envoyer le fichier en binaire
                            client_sockets.sendall(file_to_send_bytes)                                                       # on envoie le fichier au client dispo 
                            result_dictionary[file_to_send] = client_sockets                                                 # le path est en premier lieu associé au client
                            print(f'sending "{file_to_send}"')
                            client_state[client_sockets] = 'working'                                                         # ils ne sont plus pret
                            working_list.append(client_sockets)            
                            sleep(1)
        sleep(1)                                                                                                               # pas reelement necessaire mais c'est pour eviter un nombre trop important de boucles inutiles
    return 

    



def reception(working_list, client_state, fichiers_CAD_copy, result_dictionary, number_of_received_import):
    while True:
        if len(client_state)>0:
            if working_list != []:
                clients_reception_readable, clients_reception_writeable, client_reception_error = select(working_list, working_list, working_list, 5)
                if client_reception_error != []:
                    for client_socket in client_reception_error:
                        try:
                            print(f'error with "{client_socket}", the client will now close')
                            del client_state[client_socket]
                        except Exception:           # la cle a disparue?
                            print(f'error with "{client_socket}", the client can not be reach')         
                if clients_reception_readable != []:
                    for sleeping_clients in clients_reception_readable:
                        try:                                                                                                    # on peut deja lire leurs contenus avec readable, il faudra juste gerer les erreurs de wifi
                            result_wait_bytes = sleeping_clients.recv(1024)                                                     # on regarde si ils renvoient quelque chose
                            result_wait = result_wait_bytes.decode()
                            result_wait_float = float(result_wait)
                            for path in result_dictionary.keys():                                                               # on remplace les clients sockets par le timer recu apres l'import 
                                if result_dictionary[path] == sleeping_clients:
                                        result_dictionary[path] =  result_wait_float
                            number_of_received_import[0] +=1
                            client_state[sleeping_clients] = 'ready'    
                            working_list.remove(sleeping_clients)
                            print('the result of the import was acquired')
                            print( number_of_received_import[0],'/', len(fichiers_CAD_copy))
                        except socket.error:
                            print('Error while getting a client response')
                            client_state.pop(sleeping_clients)
                            for key in result_dictionary.keys():                                                               #si il y a eu un probleme de connexion          
                                if result_dictionary[key] == sleeping_clients:
                                    result_dictionary[key] = -2
                            number_of_received_import[0] +=1
                            working_list.remove(sleeping_clients)            # on l'enleve completement de la liste
                            sleeping_clients.close()
                        except Exception:
                            client_state.pop(sleeping_clients)
                            for key in result_dictionary.keys():                                                               #si il y a eu un probleme de connexion          
                                if result_dictionary[key] == sleeping_clients:
                                    result_dictionary[key] = -2
                            number_of_received_import[0] +=1
                            working_list.remove(sleeping_clients)            # on l'enleve completement de la liste
                            sleeping_clients.close()
                            
    return

def threading_in_progress(serversocket, fichiers_CAD, working_list, client_state, save_id_workspace, fichiers_CAD_copy, result_dictionary, number_of_received_import, new_clients_counter, mapping_username, mapping_password, CLI_version):
    
    accept_thread = Thread(target = accept_connexions, args=(serversocket, fichiers_CAD_copy, client_state, save_id_workspace, new_clients_counter, fichiers_CAD_copy, mapping_username, mapping_password, CLI_version), daemon = True)          #on verifie en permanence si il y a une adresse dispo
    accept_thread.start()
    
    send_thread = Thread(target = computer_in_work, args= (fichiers_CAD , working_list, client_state, result_dictionary), daemon = True)
    send_thread.start()    
    
    reception_thread = Thread(target = reception , args= (working_list, client_state, fichiers_CAD_copy, result_dictionary, number_of_received_import), daemon = True)
    reception_thread.start()
    
    return accept_thread,send_thread, reception_thread
    

def closing_thread(accept_thread, send_thread,reception_thread):
    accept_thread.join(0)    
    send_thread.join(0)
    reception_thread.join(0)
    sleep(5)    
    return
    
def closing_clients(client_state):
    for kill_clients in client_state.keys():             # pour fermer tout les clients
        closing_message = 'close'                        
        closing_message_byte= closing_message.encode('utf-8')           #pour pouvoir l'envoyer
        kill_clients.sendall(closing_message_byte)      
    return
     
  
def find_config_files_path():
   #le config file est toujours au meme endroit 
   if not os.path.exists("C:\ProgramData\cli_automation_importer"):         # destiné a etre supprime avec l'installeur
       print('do you have your config file?')
       return ''
   else:
       config_file_path = "C:\ProgramData\cli_automation_importer\cad_importer_config_file.json"
   return config_file_path

def main():  
    
    # informations sur les arguments de la fonction
    
    parser = argparse.ArgumentParser(description = 'Process the three arguments')                     # va gerer les differents arguments
    parser.add_argument('--excel_filename', type = str, help = 'name of your excel file. If not specified, the results will be put in a file named "results_import_CAD"')  #le nom du fichier excel
    parser.add_argument('--rep', type = str, help = 'name of your repertory containing the CAD files.', required= True)
    parser.add_argument('--json_ip', type = str, help = 'name of your json containing the ip adresses of the clients when the program ends.')
    args = parser.parse_args()  # parcourir les differents arguments
    
    excel_filename = args.excel_filename if args.excel_filename else 'results_import_CAD.xlsx'              # nom du fichier excel
    CAD_repertory =  args.rep                                                           # repertoire contenant les fichiers CAD
    JSON_file = find_config_files_path()
    
    # json_with_ip = args.json_ip if args.json_ip else 'ip.json'
    
    
    # vérification sur les arguments de la fonction
    
    if not verif_excel_filename(excel_filename):
        print('error : the extension of your excel filename must belong to the following ones : .xlsx, .xlsm, .xltx,.xltm' )
        return 
    
    # ip_list = read_json(json_with_ip)
    

    # variables nécessaires pour la partie scan et infos
    
        # variable nécessaires au mapping 
        
    mapping_username= input("username:")
    mapping_password = ask_password() 
    share_path = ''
    drive_letter = 'Z'
    
        # scan 
        
    file_list=[]
    path_CLI_local = None
    save_id_workspace=''     
    CLI_version = ''
    
    # initialisation des variables necessaires au serveur 
    
    global IDLE_TIME_OUT
    
    fichiers_CAD = []
    working_list = []
    client_state = {}
    timer = 0
    number_of_received_import = [0]           # pour compter le nombre d'import recu
    new_clients_counter = [0]
    
    
    # verif sur le mapping
    
    if not create_mapping_road(mapping_username, mapping_password ,drive_letter, share_path):
        return
    
    # infos et vérifications sur la CLI 

    path_CLI_local= get_config_files_datas(path_CLI_local, JSON_file)
    if  path_CLI_local== None:
        print("error while trying to read config file")
        return
    
    if not verif_CLI(path_CLI_local):            # on arrete le programme en cas d'erreur 
        return 
    
    CLI_version= check_CLI_version(path_CLI_local, CLI_version)
    
    # scan du dossier 
    
    if not verif_repertory(CAD_repertory):          # on verifie que notre repertoire est valide 
        return
    
    if not scan_CAD(file_list, fichiers_CAD, CAD_repertory) :
        return
    
    fichiers_CAD_copy = list(fichiers_CAD) # copie pour la condition
    result_dictionary = {path : 0 for path in fichiers_CAD}         # a la fin, pour que chaque path soit associe a son temps d'import
    
    

    # creation du workspace et vérifications 
    
    save_id_workspace= creation_workspace_deck(save_id_workspace, path_CLI_local)       # on va passer ce workspace aux clients
    
    if save_id_workspace== None  or save_id_workspace== '':                         # mais seulement si il est valable 
        print( 'error in the workspace creation')
        return
    else:
        print('your workspace id is :', save_id_workspace)
    # creation du serveur
    
    
    serversocket= socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    serversocket.bind(( socket.gethostbyname(socket.gethostname()), 3000) )        
    serversocket.listen()
    
    # for clients in ip_list:
    #     computer_wake_up(clients)
    
    

    try:
        accept_thread, send_thread, reception_thread = threading_in_progress( serversocket, fichiers_CAD, working_list, client_state, save_id_workspace, fichiers_CAD_copy, result_dictionary, number_of_received_import, new_clients_counter, mapping_username, mapping_password, CLI_version)  
    except KeyboardInterrupt:
        for clients in client_state.keys():
            clients.close()
        serversocket.close()
    
    while timer < IDLE_TIME_OUT and number_of_received_import[0] < len( fichiers_CAD_copy ):
        if len(client_state) != 0:
            timer = 0  
            sleep(2)
        else: 
            print(' there are no clients, remaining time before disconnexion',  IDLE_TIME_OUT - timer)
            sleep(5)
            timer += 5
    
    closing_thread(accept_thread, send_thread, reception_thread)
        
    closing_clients(client_state)   
    
    results_in_excel( result_dictionary, excel_filename, client_state )
    
    serversocket.close()
    
    
    return 
   
    
main()       









